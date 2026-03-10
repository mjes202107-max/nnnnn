import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def build_schedule(filename: str) -> None:
    # define the schedule data
    schedule = [
        {"時間": "09:00~09:30", "內容": "報到", "講者": "無"},
        {"時間": "09:30~09:40", "內容": "開場致詞", "講者": "王教授"},
        {"時間": "09:40~10:05", "內容": "從 MWC 2026 到 6G：AI-RAN 標準、開源與互通測試的最新進展", "講者": "劉教授"},
        {"時間": "10:05~10:30", "內容": "基於 O-RAN 開放介面的 ISAC 無線感知技術", "講者": "陳教授"},
        {"時間": "10:30~10:50", "內容": "Break", "講者": "無"},
        {"時間": "10:50~11:20", "內容": "Federated Foundational Models in AI-RAN：Practical and\nForward Looking Perspective", "講者": "教學團隊"},
        {"時間": "11:20~12:00", "內容": "O-RAN 環境與各模組化功能介紹", "講者": "教學團隊"},
        {"時間": "12:00~13:30", "內容": "Lunch", "講者": "無"},
        {"時間": "13:30~14:00", "內容": "O-RAN 開源軟體組織簡介", "講者": "教學團隊"},
        {"時間": "14:00~14:30", "內容": "O-RAN 實驗環境建置教學", "講者": "教學團隊"},
        {"時間": "14:30~14:50", "內容": "Break", "講者": "無"},
        {"時間": "14:50~15:50", "內容": "O-RAN xApps 實作建置教學", "講者": "教學團隊"},
        {"時間": "15:50~16:30", "內容": "現場討論時間", "講者": "教學團隊"},
    ]

    df = pd.DataFrame(schedule)

    # write to excel
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    # post-process with openpyxl to apply merges/formatting
    wb = load_workbook(filename)
    ws = wb.active

    # center the header row
    for col in range(1, 4):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

    max_row = ws.max_row

    # center all schedule cells (時間/內容/講者)
    for row in range(2, max_row + 1):
        for col in range(1, 4):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # merge cells for rows with special single entries
    for row in range(2, max_row + 1):
        content = ws.cell(row=row, column=2).value
        if content in ("報到", "Break", "Lunch"):
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    # merge adjacent rows with "教學團隊" as speaker
    row = 2
    while row <= max_row:
        speaker = ws.cell(row=row, column=3).value
        if speaker == "教學團隊" and row + 1 <= max_row and ws.cell(row=row + 1, column=3).value == "教學團隊":
            ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=2)
            ws.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=3)
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
            row += 2
        else:
            row += 1

    wb.save(filename)


if __name__ == "__main__":
    build_schedule("nnnnn.xlsx")
